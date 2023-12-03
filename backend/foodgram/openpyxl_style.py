from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Font, NamedStyle
)

header_font: Font = Font(
    size=13, bold=True, vertAlign='baseline'
)
header_fill: PatternFill = PatternFill(
    fill_type='solid', start_color='EDE979', end_color='EDE979'
)
header_side: Side = Side(
    border_style='medium', color='121102'
)
header_border: Border = Border(
    left=header_side, right=header_side,
    top=header_side, bottom=header_side
)
header_alignment = Alignment(
    horizontal='center', vertical='center'
)
body_font: Font = Font(
    size=12, bold=False, vertAlign='baseline'
)
body_fill: PatternFill = PatternFill(
    fill_type='solid', start_color='F3F5A7', end_color='F3F5A7'
)
body_side: Side = Side(
    border_style='thin', color='121102'
)
body_border: Border = Border(
    left=body_side, right=body_side,
    top=body_side, bottom=body_side
)
body_alignment = Alignment(
    horizontal='center', vertical='center'
)

header_style: NamedStyle = NamedStyle(name='header_style')
header_style.font = header_font
header_style.fill = header_fill
header_style.border = header_border
header_style.alignment = header_alignment

body_style: NamedStyle = NamedStyle(name='body_style')
body_style.font = body_font
body_style.fill = body_fill
body_style.border = body_border
body_style.alignment = body_alignment
